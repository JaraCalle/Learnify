import openpyxl
from openpyxl.utils import get_column_letter
from io import BytesIO
from django.http import HttpResponse
from .interfaces import ReportGenerator

class ExcelReportGenerator(ReportGenerator):
    def generate(self, courses):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Courses"

        headers = ["Title", "Difficulty", "Price"]
        ws.append(headers)

        for course in courses:
            ws.append([course.title, course.difficulty, float(course.price)])

        for i, _ in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(i)].width = 20

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
